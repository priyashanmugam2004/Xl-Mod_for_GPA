import streamlit as st
import openpyxl
import tempfile
st.title("XL mod for GPA Calc")
xlfile=st.file_uploader("Upload XL File Here", type='xlsx')
hide_st_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            header {visibility: hidden;}
            </style>
            """
st.markdown(hide_st_style, unsafe_allow_html=True)

def fillgpa(path):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    st.write("Accesssing & Analyzing the Spreadsheet")
    st.write()

    # List Initialization
    gradelist = []
    creditlist = []

    # {To Find Word "GPA"}
    row_no = 0
    # "Converting Count to Alphabet"
    alp = [' ', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
           'V', 'W', 'X', 'Y', 'Z']
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        row_no += 1
        col_no = 0
        for cell in row:
            col_no += 1
            if (cell.value == "GPA"):
                gpa_col = col_no
                headrow = row_no
                break
    gpa_pos = alp[gpa_col]
    # {To Find Word "GPA"}

    # {To Find Word "Reg.No"}
    row_no = 0
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        row_no += 1
        col_no = 0
        for cell in row:
            col_no += 1
            if (cell.value == "Reg.No.") or (cell.value == "Reg.No") or (cell.value == "RegNo.") or (
                    cell.value == "Reg.no.") or (cell.value == "Reg.no") or (cell.value == "Reg. no") or (
                    cell.value == "Reg. No") or (cell.value == "Reg. No."):
                regcol = col_no
                headrow = row_no
                break
    # {To Find Word "Reg.No"}

    # {To Find Number of Students}
    column_name = alp[regcol]
    loopcount = 0
    regcount = 0
    regstring = "21IT"
    for cell in sheet[column_name]:
        loopcount += 1
        if (cell.value is not None and regstring in cell.value):
            regcount += 1
    no_of_students = regcount
    # {To Find Number of Students}

    # {To Find Number of Subjects}
    subcount = 0
    substring = "17"
    for cell in sheet[headrow]:
        if (cell.value is not None and substring in cell.value):
            subcount += 1
    no_of_subjects = subcount
    # {To Find Number of Subjects}

    # {To Find Student Name Start Cell}
    student_cell = headrow + 1
    # {To Find Student Name Start Cell}

    # {Process of Calculating GPA and Updating it to the Spreadsheet}
    if no_of_subjects == 6:
        cell_obj = sheet[f'D{student_cell}': f'K{student_cell + no_of_students - 1}']
        credit = sheet['D7':'I7']

        for c1, c2, c3, c4, c5, c6 in credit:
            creditlist.append(c1.value)
            creditlist.append(c2.value)
            creditlist.append(c3.value)
            creditlist.append(c4.value)
            creditlist.append(c5.value)
            creditlist.append(c6.value)

        st.write("GPAs Calculating... Working On Spread Sheet")

        total_credit = sum(creditlist)

        for i1, i2, i3, i4, i5, i6 in cell_obj:
            credit_sum = 0
            gradelist.append(i1.value)
            gradelist.append(i2.value)
            gradelist.append(i3.value)
            gradelist.append(i4.value)
            gradelist.append(i5.value)
            gradelist.append(i6.value)
            row = f'{gpa_pos}{student_cell}'

            for sub_grade in range(0, 6):
                if (gradelist[sub_grade] == "O"):
                    gradepoint = 10
                elif (gradelist[sub_grade] == "A+"):
                    gradepoint = 9
                elif (gradelist[sub_grade] == "A"):
                    gradepoint = 8
                elif (gradelist[sub_grade] == "B+"):
                    gradepoint = 7
                elif (gradelist[sub_grade] == "B"):
                    gradepoint = 6
                elif (gradelist[sub_grade] == "A"):
                    gradepoint = 5
                elif (gradelist[sub_grade] == "RA" or gradelist[sub_grade] == "AB"):
                    gradepoint = 0
                credit_sum += gradepoint * creditlist[sub_grade]
                gpa = credit_sum / total_credit
                gpa_final = round(gpa, 3)
                sheet[row] = gpa_final
            workbook.save(path)
            gradelist.clear()
            student_cell += 1

    if no_of_subjects == 7:
        cell_obj = sheet[f'D{student_cell}': f'K{student_cell + no_of_students - 1}']
        credit = sheet['D7':'J7']

        for c1, c2, c3, c4, c5, c6, c7 in credit:
            creditlist.append(c1.value)
            creditlist.append(c2.value)
            creditlist.append(c3.value)
            creditlist.append(c4.value)
            creditlist.append(c5.value)
            creditlist.append(c6.value)
            creditlist.append(c7.value)
        st.write("GPAs Calculating... Working On Spread Sheet")

        total_credit = sum(creditlist)

        for i1, i2, i3, i4, i5, i6, i7 in cell_obj:
            credit_sum = 0
            gradelist.append(i1.value)
            gradelist.append(i2.value)
            gradelist.append(i3.value)
            gradelist.append(i4.value)
            gradelist.append(i5.value)
            gradelist.append(i6.value)
            gradelist.append(i7.value)
            row = f'{gpa_pos}{student_cell}'

            for sub_grade in range(0, 7):
                if (gradelist[sub_grade] == "O"):
                    gradepoint = 10
                elif (gradelist[sub_grade] == "A+"):
                    gradepoint = 9
                elif (gradelist[sub_grade] == "A"):
                    gradepoint = 8
                elif (gradelist[sub_grade] == "B+"):
                    gradepoint = 7
                elif (gradelist[sub_grade] == "B"):
                    gradepoint = 6
                elif (gradelist[sub_grade] == "A"):
                    gradepoint = 5
                elif (gradelist[sub_grade] == "RA" or gradelist[sub_grade] == "AB"):
                    gradepoint = 0
                credit_sum += gradepoint * creditlist[sub_grade]
                gpa = credit_sum / total_credit
                gpa_final = round(gpa, 3)
                sheet[row] = gpa_final
            workbook.save(path)
            gradelist.clear()
            student_cell += 1

    if no_of_subjects == 8:
        cell_obj = sheet[f'D{student_cell}': f'K{student_cell + no_of_students - 1}']
        credit = sheet['D7':'K7']

        for c1, c2, c3, c4, c5, c6, c7, c8 in credit:
            creditlist.append(c1.value)
            creditlist.append(c2.value)
            creditlist.append(c3.value)
            creditlist.append(c4.value)
            creditlist.append(c5.value)
            creditlist.append(c6.value)
            creditlist.append(c7.value)
            creditlist.append(c8.value)
        st.write("GPAs Calculating... Working On Spread Sheet")

        total_credit = sum(creditlist)

        for i1, i2, i3, i4, i5, i6, i7, i8 in cell_obj:
            credit_sum = 0
            gradelist.append(i1.value)
            gradelist.append(i2.value)
            gradelist.append(i3.value)
            gradelist.append(i4.value)
            gradelist.append(i5.value)
            gradelist.append(i6.value)
            gradelist.append(i7.value)
            gradelist.append(i8.value)
            row = f'{gpa_pos}{student_cell}'

            for sub_grade in range(0, 8):
                if (gradelist[sub_grade] == "O"):
                    gradepoint = 10
                elif (gradelist[sub_grade] == "A+"):
                    gradepoint = 9
                elif (gradelist[sub_grade] == "A"):
                    gradepoint = 8
                elif (gradelist[sub_grade] == "B+"):
                    gradepoint = 7
                elif (gradelist[sub_grade] == "B"):
                    gradepoint = 6
                elif (gradelist[sub_grade] == "A"):
                    gradepoint = 5
                elif (gradelist[sub_grade] == "RA" or gradelist[sub_grade] == "AB"):
                    gradepoint = 0
                credit_sum += gradepoint * creditlist[sub_grade]
                gpa = credit_sum / total_credit
                gpa_final = round(gpa, 3)
                sheet[row] = gpa_final
            workbook.save(path)
            gradelist.clear()
            student_cell += 1

    if no_of_subjects == 9:
        cell_obj = sheet[f'D{student_cell}': f'K{student_cell + no_of_students - 1}']
        credit = sheet['D7':'L7']

        for c1, c2, c3, c4, c5, c6, c7, c8, c9 in credit:
            creditlist.append(c1.value)
            creditlist.append(c2.value)
            creditlist.append(c3.value)
            creditlist.append(c4.value)
            creditlist.append(c5.value)
            creditlist.append(c6.value)
            creditlist.append(c7.value)
            creditlist.append(c8.value)
            creditlist.append(c9.value)
        st.write("GPAs Calculating... Working On Spread Sheet")

        total_credit = sum(creditlist)

        for i1, i2, i3, i4, i5, i6, i7, i8, i9 in cell_obj:
            credit_sum = 0
            gradelist.append(i1.value)
            gradelist.append(i2.value)
            gradelist.append(i3.value)
            gradelist.append(i4.value)
            gradelist.append(i5.value)
            gradelist.append(i6.value)
            gradelist.append(i7.value)
            gradelist.append(i8.value)
            gradelist.append(i9.value)
            row = f'{gpa_pos}{student_cell}'

            for sub_grade in range(0, 9):
                if (gradelist[sub_grade] == "O"):
                    gradepoint = 10
                elif (gradelist[sub_grade] == "A+"):
                    gradepoint = 9
                elif (gradelist[sub_grade] == "A"):
                    gradepoint = 8
                elif (gradelist[sub_grade] == "B+"):
                    gradepoint = 7
                elif (gradelist[sub_grade] == "B"):
                    gradepoint = 6
                elif (gradelist[sub_grade] == "A"):
                    gradepoint = 5
                elif (gradelist[sub_grade] == "RA" or gradelist[sub_grade] == "AB"):
                    gradepoint = 0
                credit_sum += gradepoint * creditlist[sub_grade]
                gpa = credit_sum / total_credit
                gpa_final = round(gpa, 3)
                sheet[row] = gpa_final
            workbook.save(path)
            gradelist.clear()
            student_cell += 1

    st.success("Spreadsheet Updated!")
    workbook.save(temp_path := tempfile.NamedTemporaryFile(delete=False).name)
    return temp_path

if xlfile:
    if st.button("Modify Spreadsheet"):
        modified_workbook_path = fillgpa(xlfile)
        st.write("Download Modified Spreadsheet Here")
        st.download_button(
            "Download Modified File",
            open(modified_workbook_path, "rb").read(),
            "modified_workbook.xlsx",
            "xlsx",
            key='download_xlsx'
        )
